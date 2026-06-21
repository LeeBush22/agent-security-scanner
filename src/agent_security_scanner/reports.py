from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any

from agent_security_scanner.i18n import (
    Language,
    category_label,
    effort_label,
    remediation_steps,
    remediation_summary,
    rule_description,
    rule_recommendation,
    rule_title,
    severity_label,
)
from agent_security_scanner.models import ScanResult


PDF_THEME_BLUE = "#2F5597"
PDF_DARK_TEXT = "#1F2937"
PDF_BORDER = "#B8C2CC"
PDF_ROW_ALT = "#F3F6FA"


def write_excel_report(result: ScanResult, output_path: Path, language: Language = Language.EN) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = _label("Summary", "摘要", language)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    summary.append([_label("Agent Security Scanner Report", "Agent Security Scanner 扫描报告", language)])
    summary["A1"].font = Font(size=16, bold=True)
    summary.append([_label("Target", "扫描目标", language), result.target])
    summary.append([_label("Total findings", "发现总数", language), result.summary.total])
    summary.append([])
    summary.append([_label("Severity", "风险等级", language), _label("Count", "数量", language)])
    for severity, count in result.summary.by_severity.items():
        summary.append([severity_label(severity, language), count])
    summary.append([])
    summary.append([_label("Category", "类型", language), _label("Count", "数量", language)])
    for category, count in result.summary.by_category.items():
        summary.append([category_label(category, language), count])

    findings = workbook.create_sheet(_label("Findings", "发现", language))
    findings_headers = [
        _label("Severity", "风险等级", language),
        _label("Rule", "规则", language),
        _label("Category", "类型", language),
        _label("File", "文件", language),
        _label("Line", "行号", language),
        _label("Title", "标题", language),
        _label("Evidence", "证据", language),
        _label("Recommendation", "建议", language),
        _label("Fingerprint", "指纹", language),
    ]
    findings.append(findings_headers)
    for finding in result.findings:
        findings.append(
            [
                severity_label(finding.severity.value, language),
                finding.rule_id,
                category_label(finding.category.value, language),
                finding.file_path,
                finding.line,
                rule_title(finding.rule_id, finding.title, language),
                finding.evidence,
                rule_recommendation(finding.rule_id, finding.recommendation, language),
                finding.fingerprint,
            ]
        )

    remediation = workbook.create_sheet(_label("Remediation", "修复", language))
    remediation.append(
        [
            _label("Rule", "规则", language),
            _label("File", "文件", language),
            _label("Effort", "修复难度", language),
            _label("Automatable", "可自动化", language),
            _label("Summary", "摘要", language),
            _label("Steps", "步骤", language),
        ]
    )
    for finding in result.findings:
        remediation_summary_text = ""
        remediation_steps_text = ""
        if finding.remediation:
            remediation_summary_text = remediation_summary(finding.rule_id, finding.remediation.summary, language)
            remediation_steps_text = "\n".join(remediation_steps(finding.rule_id, finding.remediation.steps, language))
        remediation.append(
            [
                finding.rule_id,
                finding.file_path,
                effort_label(finding.remediation.effort, language) if finding.remediation else "",
                _bool_label(finding.remediation.automatable, language) if finding.remediation else "",
                remediation_summary_text,
                remediation_steps_text,
            ]
        )

    rules = workbook.create_sheet(_label("Rules", "规则", language))
    rules.append(
        [
            _label("Rule", "规则", language),
            _label("Title", "标题", language),
            _label("Severity", "风险等级", language),
            _label("Category", "类型", language),
            _label("Description", "说明", language),
        ]
    )
    seen_rules: set[str] = set()
    for finding in result.findings:
        if finding.rule_id in seen_rules:
            continue
        seen_rules.add(finding.rule_id)
        rules.append(
            [
                finding.rule_id,
                rule_title(finding.rule_id, finding.title, language),
                severity_label(finding.severity.value, language),
                category_label(finding.category.value, language),
                rule_description(finding.rule_id, finding.description, language),
            ]
        )

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        _autosize(sheet)
        if sheet.title != summary.title and sheet.max_row >= 2 and sheet.max_column >= 2:
            table_ref = f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
            table = Table(displayName=f"Sheet{workbook.worksheets.index(sheet) + 1}Table", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

    workbook.save(output_path)
    return output_path


def write_pdf_report(result: ScanResult, output_path: Path, language: Language = Language.EN) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, TableStyle

    font_name = "Helvetica"
    bold_font_name = "Helvetica-Bold"
    needs_cjk_font = language == Language.ZH or _result_contains_cjk(result)
    if needs_cjk_font:
        font_name = "STSong-Light"
        bold_font_name = "STSong-Light"
        try:
            pdfmetrics.registerFont(UnicodeCIDFont(font_name))
        except Exception:
            font_name = "Helvetica"
            bold_font_name = "Helvetica-Bold"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        title=_label("Agent Security Scanner Report", "Agent Security Scanner 扫描报告", language),
        leftMargin=22 * mm,
        rightMargin=22 * mm,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Heading2", "Heading3", "Normal", "Italic"):
        styles[style_name].fontName = font_name
    styles["Title"].fontName = bold_font_name
    styles["Heading2"].fontName = bold_font_name
    styles["Heading3"].fontName = bold_font_name
    styles["Title"].textColor = colors.HexColor(PDF_THEME_BLUE)
    styles["Title"].fontSize = 20
    styles["Title"].leading = 26
    styles["Title"].alignment = 1
    styles["Heading2"].textColor = colors.HexColor(PDF_THEME_BLUE)
    styles["Heading2"].fontSize = 14
    styles["Heading2"].leading = 20
    styles["Heading3"].textColor = colors.HexColor(PDF_DARK_TEXT)
    styles["Heading3"].fontSize = 11
    styles["Heading3"].leading = 15
    styles["Normal"].fontSize = 9.5
    styles["Normal"].leading = 15
    styles["Italic"].fontSize = 9
    styles["Italic"].leading = 14
    cell_style = ParagraphStyle(
        "AgentScanCell",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor(PDF_DARK_TEXT),
    )

    story = [
        Paragraph(_text("Agent Security Scanner Report", "Agent Security Scanner 扫描报告", language), styles["Title"]),
        Spacer(1, 18),
        Paragraph(_text("1. Report Overview", "一、报告概览", language), styles["Heading2"]),
        Paragraph(f"{_text('Target', '项目路径', language)}: {_text(result.target, result.target, language)}", styles["Normal"]),
        Paragraph(
            f"{_text('Generated at', '生成时间', language)}: {_text(_generated_at(), _generated_at(), language)}",
            styles["Normal"],
        ),
        Paragraph(f"{_text('Total findings', '发现总数', language)}: {result.summary.total}", styles["Normal"]),
        Spacer(1, 8),
    ]

    summary_rows = [
        [_text("Dimension", "统计维度", language), _text("Item", "项目", language), _text("Count", "数量", language)]
    ]
    summary_rows.extend(
        [
            [_text("Severity", "风险等级", language), severity_label(severity, language), str(count)]
            for severity, count in result.summary.by_severity.items()
        ]
    )
    summary_rows.extend(
        [
            [_text("Category", "风险类型", language), category_label(category, language), str(count)]
            for category, count in result.summary.by_category.items()
        ]
    )
    story.append(
        _pdf_table(
            summary_rows,
            font_name=font_name,
            bold_font_name=bold_font_name,
            col_widths=[90, 300, 55],
            cell_style=cell_style,
        )
    )
    story.append(Spacer(1, 14))
    story.append(Paragraph(_text("2. Findings", "二、风险发现", language), styles["Heading2"]))

    if not result.findings:
        story.append(Paragraph(_text("No findings.", "未发现风险。", language), styles["Normal"]))
    else:
        overview_rows = [
            [
                _text("Rule", "规则", language),
                _text("Severity", "风险等级", language),
                _text("Category", "类型", language),
                _text("File", "文件", language),
                _text("Line", "行号", language),
                _text("Title", "标题", language),
            ]
        ]
        for finding in result.findings[:30]:
            overview_rows.append(
                [
                    finding.rule_id,
                    _severity_pdf_label(finding.severity.value, language),
                    category_label(finding.category.value, language),
                    finding.file_path,
                    "" if finding.line is None else str(finding.line),
                    rule_title(finding.rule_id, finding.title, language),
                ]
            )
        story.append(
            _pdf_table(
                overview_rows,
                font_name=font_name,
                bold_font_name=bold_font_name,
                col_widths=[40, 50, 65, 130, 32, 128],
                cell_style=cell_style,
            )
        )
        story.append(Spacer(1, 14))

    story.append(Paragraph(_text("3. Remediation Guidance", "三、修复建议", language), styles["Heading2"]))
    for index, finding in enumerate(result.findings, start=1):
        location = finding.file_path if not finding.line else f"{finding.file_path}:{finding.line}"
        title = f"{index}. {finding.rule_id}: {rule_title(finding.rule_id, finding.title, language)}"
        if language == Language.EN:
            title = f"{index}. {finding.rule_id}: {finding.title}"
        story.append(Paragraph(_xml(title), styles["Heading3"]))
        detail_rows = [
            [_text("Severity", "风险等级", language), _severity_pdf_label(finding.severity.value, language)],
            [_text("Category", "类型", language), category_label(finding.category.value, language)],
            [_text("File", "文件", language), location],
            [_text("Evidence", "证据", language), finding.evidence or ""],
            [
                _text("Recommendation", "建议", language),
                rule_recommendation(finding.rule_id, finding.recommendation, language),
            ],
        ]
        if finding.remediation:
            detail_rows.extend(
                [
                    [
                        _text("Remediation summary", "修复摘要", language),
                        remediation_summary(finding.rule_id, finding.remediation.summary, language),
                    ],
                    [
                        _text("Remediation steps", "修复步骤", language),
                        _numbered_steps(remediation_steps(finding.rule_id, finding.remediation.steps, language)),
                    ],
                ]
            )
        story.append(
            _pdf_key_value_table(
                detail_rows,
                font_name=font_name,
                bold_font_name=bold_font_name,
                cell_style=cell_style,
                language=language,
            )
        )
        story.append(Spacer(1, 8))

    doc.build(story, onFirstPage=_pdf_footer(font_name, language), onLaterPages=_pdf_footer(font_name, language))
    return output_path


def _pdf_table(
    rows: list[list[str]],
    font_name: str = "Helvetica",
    bold_font_name: str = "Helvetica-Bold",
    col_widths: list[int] | None = None,
    cell_style=None,
) -> Any:
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import Table as PdfTable, TableStyle

    prepared = _pdf_table_rows(rows, cell_style)
    table = PdfTable(prepared, hAlign="LEFT", colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PDF_THEME_BLUE)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTNAME", (0, 1), (-1, -1), font_name),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(PDF_BORDER)),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(PDF_ROW_ALT)]),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4 * mm),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4 * mm),
            ]
        )
    )
    return table


def _pdf_key_value_table(
    rows: list[list[str]],
    font_name: str = "Helvetica",
    bold_font_name: str = "Helvetica-Bold",
    cell_style=None,
    language: Language = Language.EN,
) -> Any:
    from reportlab.lib import colors
    from reportlab.platypus import Table as PdfTable, TableStyle

    col_widths = [132, 313] if language == Language.EN else [88, 357]
    table = PdfTable(_pdf_key_value_rows(rows, cell_style), hAlign="LEFT", colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(PDF_ROW_ALT)),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor(PDF_THEME_BLUE)),
                ("FONTNAME", (0, 0), (0, -1), bold_font_name),
                ("FONTNAME", (1, 0), (1, -1), font_name),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(PDF_BORDER)),
                ("PADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def _pdf_table_rows(rows: list[list[str]], cell_style=None) -> list[list[object]]:
    from reportlab.platypus import Paragraph

    if cell_style is None:
        return [[_xml(value) for value in row] for row in rows]
    header = [[_xml(value) for value in rows[0]]]
    body = [[_pdf_cell(value, cell_style) for value in row] for row in rows[1:]]
    return header + body


def _pdf_key_value_rows(rows: list[list[str]], cell_style=None) -> list[list[object]]:
    if cell_style is None:
        return [[_xml(value) for value in row] for row in rows]
    return [[_xml(row[0]), _pdf_cell(row[1], cell_style)] for row in rows]


def _pdf_cell(value: str, cell_style):
    from reportlab.platypus import Paragraph

    if _severity_color(value):
        return Paragraph(_severity_markup(value), cell_style)
    return Paragraph(_paragraph_markup(value), cell_style)


def _paragraph_markup(value: str) -> str:
    return _xml(value).replace("&lt;br/&gt;", "<br/>")


def _severity_markup(value: str) -> str:
    color = _severity_color(value) or PDF_DARK_TEXT
    return f'<font color="{color}"><b>{_xml(value)}</b></font>'


def _severity_color(value: str) -> str | None:
    normalized = value.strip().lower()
    return {
        "critical": "#B91C1C",
        "严重": "#B91C1C",
        "high": "#C2410C",
        "高危": "#C2410C",
        "medium": "#A16207",
        "中危": "#A16207",
        "low": "#0369A1",
        "低危": "#0369A1",
        "info": "#047857",
        "信息": "#047857",
    }.get(normalized)


def _generated_at() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _severity_pdf_label(value: str, language: Language) -> str:
    label = severity_label(value, language)
    if language == Language.ZH:
        marker = {
            "critical": "严重",
            "high": "高危",
            "medium": "中危",
            "low": "低危",
            "info": "信息",
        }.get(value, label)
        return marker
    return label


def _result_contains_cjk(result: ScanResult) -> bool:
    values = [result.target]
    for finding in result.findings:
        values.extend(
            [
                finding.title,
                finding.description,
                finding.file_path,
                finding.evidence or "",
                finding.recommendation,
            ]
        )
        if finding.remediation:
            values.append(finding.remediation.summary)
            values.extend(finding.remediation.steps)
    return any(_contains_cjk(value) for value in values)


def _contains_cjk(value: str) -> bool:
    return any("\u3400" <= char <= "\u9fff" for char in value)


def _numbered_steps(steps: list[str]) -> str:
    return "<br/>".join(f"{index}. {step}" for index, step in enumerate(steps, start=1))


def _pdf_footer(font_name: str, language: Language):
    from reportlab.lib import colors

    def draw(canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont(font_name, 8)
        canvas.setFillColor(colors.HexColor("#6B7280"))
        footer = _label(
            f"Agent Security Scanner Report    Page {doc.page}",
            f"Agent Security Scanner 扫描报告    第 {doc.page} 页",
            language,
        )
        canvas.drawCentredString(doc.pagesize[0] / 2, 10 * 2.83465, footer)
        canvas.restoreState()

    return draw


def _label(english: str, chinese: str, language: Language) -> str:
    return chinese if language == Language.ZH else english


def _text(english: str, chinese: str, language: Language) -> str:
    return _xml(_label(english, chinese, language))


def _xml(value: object) -> str:
    return escape("" if value is None else str(value))


def _bool_label(value: bool, language: Language) -> str:
    if language == Language.ZH:
        return "是" if value else "否"
    return str(value).lower()


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(max((len(part) for part in value.splitlines()), default=0), 80))
        sheet.column_dimensions[column_letter].width = max(10, min(max_len + 2, 60))
