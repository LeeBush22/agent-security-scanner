from pathlib import Path

from openpyxl import load_workbook

from agent_security_scanner.i18n import Language
from agent_security_scanner.reports import write_excel_report, write_pdf_report
from agent_security_scanner.scanner import Scanner


def test_excel_and_pdf_reports_are_written(tmp_path: Path):
    (tmp_path / "app.py").write_text('key = "sk-example1234567890example1234567890"', encoding="utf-8")
    result = Scanner().scan(tmp_path)

    excel_path = write_excel_report(result, tmp_path / "report.xlsx")
    pdf_path = write_pdf_report(result, tmp_path / "report.pdf")
    zh_excel_path = write_excel_report(result, tmp_path / "report_zh.xlsx", language=Language.ZH)
    zh_pdf_path = write_pdf_report(result, tmp_path / "report_zh.pdf", language=Language.ZH)

    assert excel_path.exists()
    assert pdf_path.exists()
    assert zh_excel_path.exists()
    assert zh_pdf_path.exists()
    assert excel_path.stat().st_size > 0
    assert pdf_path.stat().st_size > 0
    assert zh_excel_path.stat().st_size > 0
    assert zh_pdf_path.stat().st_size > 0


def test_chinese_excel_remediation_is_localized(tmp_path: Path):
    (tmp_path / "app.py").write_text('key = "sk-example1234567890example1234567890"', encoding="utf-8")
    result = Scanner().scan(tmp_path)

    zh_excel_path = write_excel_report(result, tmp_path / "report_zh.xlsx", language=Language.ZH)

    workbook = load_workbook(zh_excel_path)
    remediation_sheet = workbook["修复"]
    values = "\n".join(str(cell.value or "") for row in remediation_sheet.iter_rows() for cell in row)

    assert "移除文件中的明文凭据" in values
    assert "将凭据迁移到环境变量或密钥管理系统" in values
    assert "Move the credential" not in values
    assert "Remove plaintext credentials" not in values
    assert "Review the finding" not in values


def test_english_pdf_supports_chinese_dynamic_text(tmp_path: Path):
    project_dir = tmp_path / "中文项目"
    project_dir.mkdir()
    (project_dir / "app.py").write_text('key = "sk-example1234567890example1234567890"', encoding="utf-8")
    result = Scanner().scan(project_dir)

    pdf_path = write_pdf_report(result, tmp_path / "english_report.pdf", language=Language.EN)

    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0

    try:
        from pypdf import PdfReader
    except ImportError:
        return

    text = "\n".join(page.extract_text() or "" for page in PdfReader(pdf_path).pages)
    assert "中文项目" in text
